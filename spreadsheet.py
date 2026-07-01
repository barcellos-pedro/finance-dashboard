from pathlib import Path
import re
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = BASE_DIR / "controle_gastos_cartao_agosto_2026.xlsx"
BASE_COLUMNS = ["Descrição", "Categoria", "Valor", "Cartão", "Parcelamento"]
NAT_TYPE = type(pd.NaT)


def find_input_files() -> tuple[list[Path], list[Path]]:
    csv_files: list[Path] = []
    xlsx_files: list[Path] = []

    for path in BASE_DIR.iterdir():
        if not path.is_file() or path == OUTPUT_FILE:
            continue

        suffix = path.suffix.lower()
        if suffix == ".csv":
            csv_files.append(path)
        elif suffix == ".xlsx":
            xlsx_files.append(path)

    return sorted(csv_files), sorted(xlsx_files)


def infer_final_label(path: Path) -> str:
    stem = path.stem.lower()
    if "9136" in stem:
        return "9136"
    if "1536" in stem:
        return "1536"
    return "geral"


def br_amount_to_float(value: Any) -> Optional[float]:
    if pd.isna(value):
        return None

    text = str(value).strip().replace('"', "")
    if not text:
        return None

    negative = text.startswith("-")
    if negative:
        text = text[1:].strip()

    text = text.replace(".", "").replace(",", ".")
    try:
        parsed = float(text)
        return -parsed if negative else parsed
    except ValueError:
        try:
            return float(value)
        except (TypeError, ValueError):
            return None


def excel_serial_to_date(value: Any) -> pd.Timestamp | NAT_TYPE:
    try:
        return pd.to_datetime("1899-12-30") + pd.to_timedelta(int(float(value)), unit="D")
    except (TypeError, ValueError):
        return pd.NaT


def should_ignore_description(text: str) -> bool:
    lowered = text.lower()
    return "pagamento recebido" in lowered or "pagamento com saldo" in lowered


def suggest_cat(description: Any) -> str:
    rules = [
        ("Uber|Metro", "Transporte"),
        ("Vivo|Google One|Spotify|Apple", "Assinaturas/Serviços"),
        ("Steam|Amazon|Shein|Samsung|Ticketmaster|Lola|Joalheria|Cosmeticos", "Compras"),
        ("Airbnb", "Viagem"),
        ("Zé Delivery|Duco|Espetus|Bella|Distribuidora De Carne|Milk Tea", "Alimentação"),
        ("Pix", "Pix/Transferência"),
        ("IOF", "Taxas"),
        ("Pagamento recebido|Pagamento Com Saldo", "Pagamento/Crédito"),
    ]

    for pattern, category in rules:
        if re.search(pattern, str(description), re.I):
            return category
    return "Outros"


def load_nubank_transactions(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    try:
        df = pd.read_csv(path, dtype=str)
        for _, row in df.iterrows():
            title = str(row.get("title", "")).strip()
            if should_ignore_description(title):
                continue

            amount = br_amount_to_float(row.get("amount"))
            tipo = "Pagamento/Crédito" if amount is not None and amount < 0 else "Compra"
            rows.append(
                {
                    "Data": pd.to_datetime(str(row.get("date", "")), errors="coerce"),
                    "Descrição": title,
                    "Parcelamento": "",
                    "Valor": amount,
                    "Cartão": "Nubank",
                    "Tipo": tipo,
                }
            )
    except Exception as exc:  # pragma: no cover - defensive path
        print("CSV error", exc)

    return rows


def load_itau_transactions(files: List[tuple[Path, str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for path, final in files:
        try:
            raw = pd.read_excel(path, header=None, engine="openpyxl")
            header_idx = None
            for index in range(len(raw)):
                values = [str(value).strip()
                          for value in raw.iloc[index].tolist()]
                if "Data" in values and "Lançamento" in values and "Valor" in values:
                    header_idx = index
                    break

            if header_idx is None:
                continue

            headers = [str(value).strip() if not pd.isna(value)
                       else "" for value in raw.iloc[header_idx].tolist()]
            data = raw.iloc[header_idx + 1:].copy()
            data.columns = headers

            for _, row in data.iterrows():
                desc = row.get("Lançamento")
                if pd.isna(desc):
                    continue

                desc_text = str(desc).strip()
                if should_ignore_description(desc_text):
                    continue

                lowered = desc_text.lower()
                if lowered.startswith("subtotal") or lowered.startswith("importante"):
                    continue

                val = row.get("Valor")
                amount = float(val) if not pd.isna(val) else None
                if amount is None:
                    continue

                dt = excel_serial_to_date(row.get("Data"))
                tipo = "Pagamento/Crédito" if amount < 0 else "Compra"
                card_name = "Itaú Click" if final == "9136" else "Itaú Uniclass" if final == "1536" else "Itaú"
                rows.append(
                    {
                        "Data": dt,
                        "Descrição": desc_text,
                        "Parcelamento": "" if pd.isna(row.get("Parcelamento")) else str(row.get("Parcelamento")).strip(),
                        "Valor": amount,
                        "Cartão": card_name,
                        "Tipo": tipo,
                    }
                )
        except Exception as exc:  # pragma: no cover - defensive path
            print("xlsx error", path.name, exc)

    return rows


def build_transactions() -> pd.DataFrame:
    csv_files, xlsx_files = find_input_files()
    rows: List[Dict[str, Any]] = []

    for csv_file in csv_files:
        rows.extend(load_nubank_transactions(csv_file))

    itau_files = [(path, infer_final_label(path)) for path in xlsx_files]
    rows.extend(load_itau_transactions(itau_files))

    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["Data", "Descrição", "Categoria", "Valor", "Cartão", "Parcelamento"])

    df = df.sort_values(["Data", "Cartão", "Descrição"],
                        na_position="last").reset_index(drop=True)
    df["Categoria"] = df["Descrição"].apply(suggest_cat)
    df["Parcelamento"] = df["Parcelamento"].fillna("")

    return df[["Data", "Descrição", "Categoria", "Valor", "Cartão", "Parcelamento"]].copy()


def style_worksheet(ws, headers: List[str]) -> None:
    column_index = {header: index + 1 for index, header in enumerate(headers)}
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = Border(bottom=thin)
            if cell.column == column_index["Valor"]:
                cell.number_format = "R$ #,##0.00;[Red]-R$ #,##0.00"


def create_workbook(df: pd.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Lançamentos"

    headers = BASE_COLUMNS
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append([row[column] for column in headers])

    style_worksheet(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    end_column = get_column_letter(ws.max_column)
    table = Table(displayName="TabelaLancamentos",
                  ref=f"A1:{end_column}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)

    widths = {"A": 36, "B": 22, "C": 12, "D": 16, "E": 16, "F": 18}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    dv_yesno = DataValidation(
        type="list", formula1='"Sim,Não"', allow_blank=False)
    ws.add_data_validation(dv_yesno)
    dv_yesno.add(f"F2:F{ws.max_row}")

    res = wb.create_sheet("Resumo")
    assert res is not None
    res["A1"] = "Resumo dos gastos"
    res["A1"].font = Font(size=16, bold=True, color="1F4E78")

    summary_items = [
        ("Total de compras", '=SUMIF(TabelaLancamentos[Valor],">0")'),
        ("Pagamentos/créditos", '=SUMIF(TabelaLancamentos[Valor],"<0")'),
        ("Saldo líquido", "=SUM(TabelaLancamentos[Valor])"),
    ]
    for index, (label, formula) in enumerate(summary_items, start=3):
        res[f"A{index}"] = label
        res[f"B{index}"] = formula
        res[f"B{index}"].number_format = "R$ #,##0.00;[Red]-R$ #,##0.00"

    res["D2"] = "Por categoria"
    res["D2"].font = Font(bold=True)
    categories = sorted(df["Categoria"].dropna().unique())
    res["D3"] = "Categoria"
    res["E3"] = "Total"

    for row_number, category in enumerate(categories, start=4):
        res[f"D{row_number}"] = category
        res[f"E{row_number}"] = f"=SUMIF(TabelaLancamentos[Categoria],D{row_number},TabelaLancamentos[Valor])"
        res[f"E{row_number}"].number_format = "R$ #,##0.00;[Red]-R$ #,##0.00"

    for cell in res[3]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for column in ["A", "D"]:
        res.column_dimensions[column].width = 26
    for column in ["B", "E"]:
        res.column_dimensions[column].width = 16

    pie = PieChart()
    pie.title = "Gastos por categoria"
    labels = Reference(res, min_col=4, min_row=4, max_row=3 + len(categories))
    data = Reference(res, min_col=5, min_row=3, max_row=3 + len(categories))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    res.add_chart(pie, "G3")

    inst = wb.create_sheet("Como usar")
    assert inst is not None
    inst.append(["Como usar a planilha"])
    inst["A1"].font = Font(size=16, bold=True, color="1F4E78")
    lines = [
        "1. Na aba Lançamentos, revise a Categoria se quiser ajustar a classificação.",
        "4. A aba Resumo consolida total de compras, créditos, saldo líquido e quantidade de itens marcados para dividir.",
    ]
    for index, line in enumerate(lines, start=3):
        inst[f"A{index}"] = line
    inst.column_dimensions["A"].width = 110

    wb.save(output_path)


def main() -> None:
    df = build_transactions()
    create_workbook(df, OUTPUT_FILE)
    print(OUTPUT_FILE.name, len(df),
          df["Valor"].sum(), df[df["Valor"] > 0]["Valor"].sum())


if __name__ == "__main__":
    main()
